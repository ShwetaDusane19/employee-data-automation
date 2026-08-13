import pandas as pd
from sqlalchemy import create_engine
from urllib.parse import quote_plus
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils.dataframe import dataframe_to_rows

# --- Database Credentials ---
DB_USER = "root"          # Replace with your MySQL username
DB_PASSWORD = quote_plus("Sql@2026")  # Replace with your MySQL password
DB_HOST = "localhost"
DB_PORT = "3306"
DB_NAME = "hr_analytics"

def build_excel_report():
    print("📊 Querying MySQL and generating management report...")
    
    # 1. Establish DB Connection
    connection_str = f"mysql+pymysql://{DB_USER}:{DB_PASSWORD}@{DB_HOST}:{DB_PORT}/{DB_NAME}"
    engine = create_engine(connection_str)

    # 2. Extract Data via SQL Queries
    # Query A: Department-level Metrics
    dept_query = """
    SELECT 
        department,
        COUNT(emp_id) AS total_employees,
        ROUND(AVG(base_salary), 2) AS avg_base_salary,
        ROUND(SUM(base_salary), 2) AS total_payroll
    FROM clean_employees
    GROUP BY department
    ORDER BY total_payroll DESC;
    """
    df_dept = pd.read_sql(dept_query, con=engine)

    # Query B: Full Clean Employee List
    clean_query = "SELECT emp_id, full_name, email, department, joining_date, base_salary FROM clean_employees;"
    df_clean = pd.read_sql(clean_query, con=engine)

    # Query C: Data Exceptions Log
    exception_query = "SELECT exception_id, emp_id, first_name, last_name, email, department, issue_reason FROM data_exceptions;"
    df_exceptions = pd.read_sql(exception_query, con=engine)

    # 3. Create Workbook
    wb = openpyxl.Workbook()
    
    # --- Tab 1: Executive Summary ---
    ws1 = wb.active
    ws1.title = "Executive Summary"
    
    # Styling definitions
    header_fill = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid") # Dark Blue
    header_font = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
    title_font = Font(name="Calibri", size=16, bold=True, color="1F4E78")
    bold_font = Font(name="Calibri", size=11, bold=True)
    
    # Add Title Block
    ws1.append(["HR Analytics & Exception Summary Report"])
    ws1.cell(row=1, column=1).font = title_font
    ws1.append([]) # Empty row

    # Write KPI Aggregations
    ws1.append(["Key Metrics", "Value"])
    ws1.append(["Total Clean Records", len(df_clean)])
    ws1.append(["Total Flagged Exceptions", len(df_exceptions)])
    ws1.append(["Total Annual Payroll Incurred", df_dept['total_payroll'].sum()])
    
    for r in range(3, 7):
        ws1.cell(row=r, column=1).font = bold_font

    ws1.append([]) # Empty row
    ws1.append(["Department Breakdowns"])
    ws1.cell(row=8, column=1).font = title_font
    ws1.append([])

    # Write Dept Summary Table
    for r in dataframe_to_rows(df_dept, index=False, header=True):
        ws1.append(r)

    # Style Table Headers in Tab 1
    header_row_idx = 10
    for cell in ws1[header_row_idx]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center")

    # --- Tab 2: Clean Employee Roster ---
    ws2 = wb.create_sheet(title="Clean Employees")
    for r in dataframe_to_rows(df_clean, index=False, header=True):
        ws2.append(r)
    for cell in ws2[1]:
        cell.fill = header_fill
        cell.font = header_font

    # --- Tab 3: Exception Log ---
    ws3 = wb.create_sheet(title="Exception Records")
    for r in dataframe_to_rows(df_exceptions, index=False, header=True):
        ws3.append(r)
    
    exception_fill = PatternFill(start_color="C00000", end_color="C00000", fill_type="solid") # Dark Red
    for cell in ws3[1]:
        cell.fill = exception_fill
        cell.font = header_font

    # Auto-adjust column widths across all sheets
    for ws in [ws1, ws2, ws3]:
        for col in ws.columns:
            max_len = max(len(str(cell.value or '')) for cell in col)
            col_letter = openpyxl.utils.get_column_letter(col[0].column)
            ws.column_dimensions[col_letter].width = max(max_len + 3, 12)

    # 4. Save Final Report
    output_report = "data/output/HR_Management_Report.xlsx"
    wb.save(output_report)
    print(f"🎉 Report successfully created at: {output_report}")

if __name__ == "__main__":
    build_excel_report()